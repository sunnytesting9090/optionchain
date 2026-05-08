from __future__ import annotations

import argparse
import os
import queue
import shutil
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import xlwings as xw
from kiteconnect import KiteConnect, KiteTicker
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "simple_option_chain.xlsx"
ACCESS_TOKEN_PATH = BASE_DIR / "access_token.txt"
CACHE_FILES = {
    "NFO": BASE_DIR / "instruments_nfo.csv",
    "BFO": BASE_DIR / "instruments_bfo.csv",
}

WORKBOOK_VERSION = "simple-option-chain-v4"
AUTOSAVE_SECONDS = 60
BATCH_RENDER_SECONDS = 0.3
QUEUE_POLL_SECONDS = 0.05
OPTION_BATCH_SIZE = 450
CACHE_MAX_AGE_HOURS = 12
DATA_START_ROW = 8
MAX_CLEAR_ROWS = 4000
DISPLAY_STRIKE_SIDE = 25
DISPLAY_STRIKE_COUNT = (DISPLAY_STRIKE_SIDE * 2) + 1
DISPLAY_END_ROW = DATA_START_ROW + DISPLAY_STRIKE_COUNT - 1

SUPPORTED_INDICES = {
    "NIFTY": {
        "name": "NIFTY",
        "segment": "NFO-OPT",
        "exchange": "NFO",
        "spot_symbol": "NSE:NIFTY 50",
    },
    "BANKNIFTY": {
        "name": "BANKNIFTY",
        "segment": "NFO-OPT",
        "exchange": "NFO",
        "spot_symbol": "NSE:NIFTY BANK",
    },
    "SENSEX": {
        "name": "SENSEX",
        "segment": "BFO-OPT",
        "exchange": "BFO",
        "spot_symbol": "BSE:SENSEX",
    },
}
INDEX_ORDER = list(SUPPORTED_INDICES.keys())
RAW_HEADERS = [
    "Index",
    "Expiry",
    "Tradingsymbol",
    "Strike",
    "Option Type",
    "Last Price",
    "Volume",
    "OI",
    "OI Day High",
    "OI Day Low",
    "Open",
    "High",
    "Low",
    "Close",
    "Best Bid Price",
    "Last Trade Time",
]
RAW_LIVE_COLUMN_START = RAW_HEADERS.index("Last Price") + 1
RAW_LIVE_COLUMN_END = RAW_HEADERS.index("Last Trade Time") + 1
RAW_LAST_PRICE_COLUMN_LETTER = get_column_letter(RAW_LIVE_COLUMN_START)


def log(message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}", flush=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Tick-based simple option chain workbook for NIFTY, BANKNIFTY, and SENSEX."
    )
    parser.add_argument("--once", action="store_true", help="Refresh the workbook once and exit.")
    parser.add_argument(
        "--rebuild-only",
        action="store_true",
        help="Create or rebuild the workbook template without opening Excel.",
    )
    return parser.parse_args()


def load_access_token() -> str:
    if not ACCESS_TOKEN_PATH.exists():
        raise FileNotFoundError(
            "access_token.txt is missing. Run run_simple_option_chain.bat and paste today's access token."
        )

    token = ACCESS_TOKEN_PATH.read_text(encoding="utf-8").strip()
    if not token:
        raise ValueError(
            "access_token.txt is empty. Run run_simple_option_chain.bat and paste today's access token."
        )

    return token


def build_kite_client(access_token: str) -> KiteConnect:
    log("Loading API credentials and access token.")
    kite = KiteConnect(api_key=config.cred["api_key"])
    kite.set_access_token(access_token)
    profile = kite.profile()
    user_name = profile.get("user_name") or profile.get("user_id") or "Unknown user"
    log(f"Connected to Kite as {user_name}.")
    return kite


def cache_is_stale(cache_path: Path) -> bool:
    if not cache_path.exists():
        return True
    age_seconds = time.time() - cache_path.stat().st_mtime
    return age_seconds > (CACHE_MAX_AGE_HOURS * 3600)


def refresh_exchange_cache(kite: KiteConnect | None, exchange: str) -> None:
    cache_path = CACHE_FILES[exchange]
    if cache_path.exists() and (kite is None or not cache_is_stale(cache_path)):
        log(f"Using cached {exchange} instruments from {cache_path.name}.")
        return

    if kite is None:
        raise FileNotFoundError(
            f"{cache_path.name} is missing and a live Kite session is not available to download it."
        )

    log(f"Refreshing the {exchange} instrument cache from Kite.")
    instruments = kite.instruments(exchange)
    pd.DataFrame(instruments).to_csv(cache_path, index=False)
    log(f"Saved {len(instruments)} rows to {cache_path.name}.")


def load_instrument_universe(kite: KiteConnect | None) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for exchange in sorted({meta["exchange"] for meta in SUPPORTED_INDICES.values()}):
        refresh_exchange_cache(kite, exchange)
        frame = pd.read_csv(CACHE_FILES[exchange], low_memory=False)
        frame["expiry"] = frame["expiry"].astype(str).str[:10]
        frame["name"] = frame["name"].astype(str)
        frame["segment"] = frame["segment"].astype(str)
        frame["instrument_type"] = frame["instrument_type"].astype(str)
        frames.append(frame)

    combined = pd.concat(frames, ignore_index=True)
    return combined[combined["segment"].isin({"NFO-OPT", "BFO-OPT"})].copy()


def build_expiry_map(instrument_frame: pd.DataFrame) -> dict[str, list[str]]:
    expiry_map: dict[str, list[str]] = {}
    for index_name, meta in SUPPORTED_INDICES.items():
        filtered = instrument_frame[
            (instrument_frame["name"] == meta["name"])
            & (instrument_frame["segment"] == meta["segment"])
            & (instrument_frame["instrument_type"].isin(["CE", "PE"]))
        ]
        expiries = sorted(filtered["expiry"].dropna().unique().tolist())
        if not expiries:
            raise RuntimeError(f"No expiries found for {index_name}.")
        expiry_map[index_name] = expiries
    return expiry_map


def nearest_expiry(expiries: list[str]) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    future = [expiry for expiry in expiries if expiry >= today]
    return future[0] if future else expiries[0]


def workbook_requires_rebuild() -> bool:
    if not WORKBOOK_PATH.exists():
        return True

    try:
        workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
        try:
            required_sheets = {"Summary", "NIFTY", "BANKNIFTY", "SENSEX", "RawData", "Meta"}
            if not required_sheets.issubset(set(workbook.sheetnames)):
                return True
            return workbook["Meta"]["A1"].value != WORKBOOK_VERSION
        finally:
            workbook.close()
    except Exception:
        return True


def backup_existing_workbook() -> None:
    if not WORKBOOK_PATH.exists():
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = WORKBOOK_PATH.with_name(f"simple_option_chain_backup_{timestamp}.xlsx")
    shutil.copy2(WORKBOOK_PATH, backup_path)
    log(f"Existing workbook backed up to {backup_path.name}.")


def style_range_border(sheet, cell_refs: list[str]) -> None:
    border = Border(
        left=Side(style="thin", color="C9C9C9"),
        right=Side(style="thin", color="C9C9C9"),
        top=Side(style="thin", color="C9C9C9"),
        bottom=Side(style="thin", color="C9C9C9"),
    )
    for cell_ref in cell_refs:
        sheet[cell_ref].border = border


def create_workbook_template(default_expiries: dict[str, str]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    summary.merge_cells("A1:F1")
    summary["A1"] = "Tick Option Chain Summary"
    summary["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    summary["A1"].alignment = Alignment(horizontal="center")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")

    summary["A3"] = "Index"
    summary["B3"] = "Nearest Expiry"
    summary["C3"] = "Spot"
    summary["D3"] = "ATM Strike"
    summary["E3"] = "ATM Straddle"
    summary["F3"] = "Status"
    for cell in summary["3:3"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="548235")
        cell.alignment = Alignment(horizontal="center")

    for row_number, index_name in enumerate(INDEX_ORDER, start=4):
        summary[f"A{row_number}"] = index_name
        summary[f"B{row_number}"] = default_expiries[index_name]
        summary[f"F{row_number}"] = "Waiting for live ticks..."

    summary["A9"] = "Last Updated"
    summary["B9"] = ""
    summary["A10"] = "Mode"
    summary["B10"] = "WebSocket tick stream + 300ms batch"
    summary["A11"] = "Autosave"
    summary["B11"] = f"{AUTOSAVE_SECONDS} seconds"
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 14
    summary.column_dimensions["E"].width = 16
    summary.column_dimensions["F"].width = 40
    style_range_border(
        summary,
        [f"{column}{row}" for row in range(3, 11) for column in ("A", "B", "C", "D", "E", "F")],
    )

    for index_name in INDEX_ORDER:
        sheet = workbook.create_sheet(index_name)
        sheet.merge_cells("A1:E1")
        sheet["A1"] = f"{index_name} Option Chain"
        sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")

        sheet["A3"] = "Nearest Expiry"
        sheet["B3"] = default_expiries[index_name]
        sheet["D3"] = "Spot"
        sheet["E3"] = ""
        sheet["A4"] = "ATM Strike"
        sheet["B4"] = ""
        sheet["D4"] = "ATM Straddle"
        sheet["E4"] = ""
        sheet["A5"] = "Last Updated"
        sheet["B5"] = ""
        sheet["D5"] = "Status"
        sheet["E5"] = "Waiting for live ticks..."
        sheet["A6"] = f"Showing ATM +/- {DISPLAY_STRIKE_SIDE} strikes"

        headers = ["ATM", "Strike", "Call LTP", "Put LTP", "Straddle Sum"]
        for column_number, header in enumerate(headers, start=1):
            cell = sheet.cell(row=7, column=column_number)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="548235")
            cell.alignment = Alignment(horizontal="center")

        style_range_border(
            sheet,
            ["A3", "B3", "D3", "E3", "A4", "B4", "D4", "E4", "A5", "B5", "D5", "E5", "A6"]
            + [f"{column}7" for column in ("A", "B", "C", "D", "E")],
        )
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 14
        sheet.column_dimensions["E"].width = 16
        sheet.freeze_panes = f"A{DATA_START_ROW}"
        sheet.sheet_view.showGridLines = False

    raw_sheet = workbook.create_sheet("RawData")
    for column_number, header in enumerate(RAW_HEADERS, start=1):
        cell = raw_sheet.cell(row=1, column=column_number)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7030A0")
        cell.alignment = Alignment(horizontal="center")
    raw_sheet.freeze_panes = "A2"
    raw_sheet.sheet_view.showGridLines = False
    raw_sheet.column_dimensions["A"].width = 14
    raw_sheet.column_dimensions["B"].width = 14
    raw_sheet.column_dimensions["C"].width = 22
    raw_sheet.column_dimensions["D"].width = 12
    raw_sheet.column_dimensions["E"].width = 12
    for column_letter in ("F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
        raw_sheet.column_dimensions[column_letter].width = 14
    raw_sheet.column_dimensions["P"].width = 22

    meta_sheet = workbook.create_sheet("Meta")
    meta_sheet["A1"] = WORKBOOK_VERSION
    meta_sheet.sheet_state = "hidden"
    workbook.save(WORKBOOK_PATH)


def ensure_workbook(default_expiries: dict[str, str]) -> None:
    if workbook_requires_rebuild():
        if WORKBOOK_PATH.exists():
            backup_existing_workbook()
        log("Creating the simple_option_chain.xlsx workbook.")
        create_workbook_template(default_expiries)
    else:
        log("Existing simple_option_chain.xlsx structure is valid, so it will be reused.")


def find_open_workbook() -> tuple[xw.App | None, xw.Book | None]:
    workbook_path = os.path.normcase(str(WORKBOOK_PATH.resolve()))
    for app in xw.apps:
        for book in app.books:
            try:
                if os.path.normcase(book.fullname) == workbook_path:
                    return app, book
            except Exception:
                continue
    return None, None


def open_excel_workbook() -> tuple[xw.App, xw.Book]:
    existing_app, existing_book = find_open_workbook()
    if existing_app and existing_book:
        log("Using the simple workbook that is already open in Excel.")
        return existing_app, existing_book

    log("Opening simple_option_chain.xlsx in Excel.")
    app = xw.App(visible=True, add_book=False)
    app.display_alerts = False
    book = app.books.open(str(WORKBOOK_PATH))
    book.activate(steal_focus=True)
    return app, book


def number_or_none(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def display_number(value: Any) -> int | float | None:
    number = number_or_none(value)
    if number is None:
        return None
    return int(number) if number.is_integer() else round(number, 2)


def format_timestamp(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def build_runtime_plan(kite: KiteConnect, instrument_frame: pd.DataFrame) -> dict[str, Any]:
    spot_symbols = [SUPPORTED_INDICES[index_name]["spot_symbol"] for index_name in INDEX_ORDER]
    spot_quotes = kite.quote(spot_symbols)

    expiry_map = build_expiry_map(instrument_frame)
    plan: dict[str, Any] = {
        "indices": {},
        "option_tokens": [],
        "spot_tokens": [],
        "subscription_tokens": [],
        "option_token_to_index": {},
        "spot_token_to_index": {},
        "token_to_raw_row": {},
        "raw_static_rows": [],
    }

    raw_row_number = 2
    for index_name in INDEX_ORDER:
        meta = SUPPORTED_INDICES[index_name]
        expiry = nearest_expiry(expiry_map[index_name])
        filtered = instrument_frame[
            (instrument_frame["name"] == meta["name"])
            & (instrument_frame["segment"] == meta["segment"])
            & (instrument_frame["expiry"] == expiry)
            & (instrument_frame["instrument_type"].isin(["CE", "PE"]))
        ].copy()

        ce_frame = (
            filtered[filtered["instrument_type"] == "CE"][["strike", "instrument_token"]]
            .drop_duplicates("strike")
            .rename(columns={"instrument_token": "ce_token"})
        )
        pe_frame = (
            filtered[filtered["instrument_type"] == "PE"][["strike", "instrument_token"]]
            .drop_duplicates("strike")
            .rename(columns={"instrument_token": "pe_token"})
        )
        paired = ce_frame.merge(pe_frame, on="strike", how="inner").sort_values("strike").reset_index(drop=True)
        if paired.empty:
            raise RuntimeError(f"No CE/PE strike pairs found for {index_name} {expiry}.")
        paired["sheet_row"] = range(DATA_START_ROW, DATA_START_ROW + len(paired))

        contracts = filtered[
            ["tradingsymbol", "instrument_token", "exchange", "segment", "strike", "instrument_type"]
        ].copy()
        contracts["instrument_token"] = contracts["instrument_token"].astype(int)
        contracts.sort_values(["strike", "instrument_type"], inplace=True, ignore_index=True)

        spot_quote = spot_quotes.get(meta["spot_symbol"], {})
        spot_token = spot_quote.get("instrument_token")
        if spot_token is None:
            raise RuntimeError(f"Spot instrument token was not returned for {meta['spot_symbol']}.")
        spot_token = int(spot_token)

        plan["spot_tokens"].append(spot_token)
        plan["spot_token_to_index"][spot_token] = index_name
        plan["subscription_tokens"].append(spot_token)

        plan["raw_static_rows"].append(
            [
                index_name,
                expiry,
                meta["spot_symbol"],
                None,
                None,
            ]
            + [None] * (len(RAW_HEADERS) - 5)
        )
        plan["token_to_raw_row"][spot_token] = raw_row_number
        raw_row_number += 1

        for contract in contracts.itertuples(index=False):
            strike_output = display_number(contract.strike)
            token = int(contract.instrument_token)
            plan["option_tokens"].append(token)
            plan["subscription_tokens"].append(token)
            plan["option_token_to_index"][token] = index_name
            plan["token_to_raw_row"][token] = raw_row_number
            plan["raw_static_rows"].append(
                [
                    index_name,
                    expiry,
                    contract.tradingsymbol,
                    strike_output,
                    contract.instrument_type,
                ]
                + [None] * (len(RAW_HEADERS) - 5)
            )
            raw_row_number += 1

        paired["ce_raw_row"] = paired["ce_token"].map(plan["token_to_raw_row"])
        paired["pe_raw_row"] = paired["pe_token"].map(plan["token_to_raw_row"])

        plan["indices"][index_name] = {
            "expiry": expiry,
            "paired": paired,
            "spot_symbol": meta["spot_symbol"],
            "spot_token": spot_token,
            "last_window_anchor": None,
        }

    return plan


def build_index_formula_rows(index_plan: dict[str, Any], start_position: int, end_position: int) -> list[list[Any]]:
    formula_rows: list[list[Any]] = []
    output_row = DATA_START_ROW
    for row in index_plan["paired"].iloc[start_position:end_position].itertuples(index=False):
        sheet_row = output_row
        strike_output = display_number(row.strike)
        ce_raw_row = int(row.ce_raw_row)
        pe_raw_row = int(row.pe_raw_row)
        formula_rows.append(
            [
                f'=IF(B{sheet_row}=$B$4,"ATM","")',
                strike_output,
                f'=IF(RawData!{RAW_LAST_PRICE_COLUMN_LETTER}{ce_raw_row}="","",RawData!{RAW_LAST_PRICE_COLUMN_LETTER}{ce_raw_row})',
                f'=IF(RawData!{RAW_LAST_PRICE_COLUMN_LETTER}{pe_raw_row}="","",RawData!{RAW_LAST_PRICE_COLUMN_LETTER}{pe_raw_row})',
                f'=IF(OR(C{sheet_row}="",D{sheet_row}=""),"",C{sheet_row}+D{sheet_row})',
            ]
        )
        output_row += 1
    return formula_rows


def build_display_window(index_plan: dict[str, Any], atm_position: int | None) -> tuple[int, int]:
    total_rows = len(index_plan["paired"])
    if total_rows == 0:
        return 0, 0

    if atm_position is None:
        atm_position = total_rows // 2

    start_position = max(0, atm_position - DISPLAY_STRIKE_SIDE)
    end_position = min(total_rows, atm_position + DISPLAY_STRIKE_SIDE + 1)

    window_size = end_position - start_position
    if window_size < DISPLAY_STRIKE_COUNT:
        pad_needed = DISPLAY_STRIKE_COUNT - window_size
        grow_left = min(start_position, pad_needed)
        start_position -= grow_left
        end_position = min(total_rows, end_position + (pad_needed - grow_left))
        if end_position - start_position < DISPLAY_STRIKE_COUNT:
            start_position = max(0, end_position - DISPLAY_STRIKE_COUNT)

    return start_position, end_position


def write_index_window(index_sheet: xw.Sheet, index_plan: dict[str, Any], atm_position: int | None) -> int | None:
    if atm_position is None:
        index_sheet.range((DATA_START_ROW, 1), (DISPLAY_END_ROW, 5)).clear_contents()
        return None

    start_position, end_position = build_display_window(index_plan, atm_position)
    formula_rows = build_index_formula_rows(index_plan, start_position, end_position)
    index_sheet.range((DATA_START_ROW, 1), (DISPLAY_END_ROW, 5)).clear_contents()
    if formula_rows:
        index_sheet.range((DATA_START_ROW, 1)).value = formula_rows
    return start_position


def prepare_workbook_for_plan(workbook: xw.Book, plan: dict[str, Any]) -> None:
    summary_sheet = workbook.sheets["Summary"]
    raw_sheet = workbook.sheets["RawData"]

    for row_number, index_name in enumerate(INDEX_ORDER, start=4):
        summary_sheet.range(f"A{row_number}").value = index_name
        summary_sheet.range(f"B{row_number}").formula = f"={index_name}!B3"
        summary_sheet.range(f"C{row_number}").formula = f"={index_name}!E3"
        summary_sheet.range(f"D{row_number}").formula = f"={index_name}!B4"
        summary_sheet.range(f"E{row_number}").formula = f"={index_name}!E4"
        summary_sheet.range(f"F{row_number}").formula = f"={index_name}!E5"

    summary_sheet.range("B9").value = ""
    summary_sheet.range("A10:B11").value = [
        ["Mode", f"WebSocket tick stream + {int(BATCH_RENDER_SECONDS * 1000)}ms batch"],
        ["Autosave", f"{AUTOSAVE_SECONDS} seconds"],
    ]

    raw_sheet.range((2, 1), (MAX_CLEAR_ROWS, len(RAW_HEADERS))).clear_contents()
    if plan["raw_static_rows"]:
        raw_sheet.range((2, 1)).value = plan["raw_static_rows"]

    for index_name in INDEX_ORDER:
        index_sheet = workbook.sheets[index_name]
        index_sheet.range("B3").value = plan["indices"][index_name]["expiry"]
        index_sheet.range("E3").value = None
        index_sheet.range("B4").value = None
        index_sheet.range("E4").value = None
        index_sheet.range("B5").value = None
        index_sheet.range("E5").value = "Waiting for live ticks..."
        index_sheet.range((DATA_START_ROW, 1), (MAX_CLEAR_ROWS, 5)).clear_contents()
        plan["indices"][index_name]["last_window_anchor"] = None

    workbook.save()


def fetch_ltp_in_batches(kite: KiteConnect, tokens: list[int]) -> dict[Any, Any]:
    quotes: dict[Any, Any] = {}
    for start_index in range(0, len(tokens), OPTION_BATCH_SIZE):
        batch = tokens[start_index : start_index + OPTION_BATCH_SIZE]
        if batch:
            quotes.update(kite.ltp(batch))
    return quotes


def seed_initial_ticks(kite: KiteConnect, plan: dict[str, Any]) -> dict[int, dict[str, Any]]:
    latest_ticks: dict[int, dict[str, Any]] = {}

    spot_symbols = [plan["indices"][index_name]["spot_symbol"] for index_name in INDEX_ORDER]
    spot_quotes = kite.quote(spot_symbols)
    for index_name in INDEX_ORDER:
        spot_symbol = plan["indices"][index_name]["spot_symbol"]
        spot_token = plan["indices"][index_name]["spot_token"]
        quote = spot_quotes.get(spot_symbol, {})
        latest_ticks[spot_token] = {
            "instrument_token": spot_token,
            "last_price": quote.get("last_price"),
            "change": quote.get("change"),
            "ohlc": quote.get("ohlc"),
            "exchange_timestamp": quote.get("timestamp") or quote.get("exchange_timestamp"),
        }

    option_quotes = fetch_ltp_in_batches(kite, plan["option_tokens"])
    for token in plan["option_tokens"]:
        quote = option_quotes.get(token) or option_quotes.get(str(token)) or {}
        latest_ticks[token] = {
            "instrument_token": token,
            "last_price": quote.get("last_price"),
        }

    return latest_ticks


def get_best_depth_entry(tick: dict[str, Any], side: str) -> dict[str, Any]:
    depth = tick.get("depth") or {}
    entries = depth.get(side) or []
    return entries[0] if entries else {}


def tick_to_raw_live_values(tick: dict[str, Any]) -> list[Any]:
    ohlc = tick.get("ohlc") or {}
    best_buy = get_best_depth_entry(tick, "buy")
    return [
        number_or_none(tick.get("last_price")),
        tick.get("volume_traded"),
        tick.get("oi"),
        tick.get("oi_day_high"),
        tick.get("oi_day_low"),
        number_or_none(ohlc.get("open")),
        number_or_none(ohlc.get("high")),
        number_or_none(ohlc.get("low")),
        number_or_none(ohlc.get("close")),
        number_or_none(best_buy.get("price")),
        format_timestamp(tick.get("last_trade_time")),
    ]


def write_raw_updates(
    raw_sheet: xw.Sheet,
    plan: dict[str, Any],
    latest_ticks: dict[int, dict[str, Any]],
    changed_tokens: set[int],
) -> None:
    updates: list[tuple[int, list[Any]]] = []
    for token in changed_tokens:
        row_number = plan["token_to_raw_row"].get(token)
        tick = latest_ticks.get(token)
        if row_number is None or tick is None:
            continue
        updates.append((row_number, tick_to_raw_live_values(tick)))

    if not updates:
        return

    updates.sort(key=lambda item: item[0])
    block_start = updates[0][0]
    block_end = updates[0][0]
    block_values: list[list[Any]] = [updates[0][1]]

    for row_number, row_values in updates[1:]:
        if row_number == block_end + 1:
            block_end = row_number
            block_values.append(row_values)
            continue

        raw_sheet.range((block_start, RAW_LIVE_COLUMN_START), (block_end, RAW_LIVE_COLUMN_END)).value = block_values
        block_start = row_number
        block_end = row_number
        block_values = [row_values]

    raw_sheet.range((block_start, RAW_LIVE_COLUMN_START), (block_end, RAW_LIVE_COLUMN_END)).value = block_values


def last_price_from_state(latest_ticks: dict[int, dict[str, Any]], token: int) -> float | None:
    tick = latest_ticks.get(token) or {}
    return number_or_none(tick.get("last_price"))


def build_index_result(
    index_name: str,
    plan: dict[str, Any],
    latest_ticks: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    index_plan = plan["indices"][index_name]
    paired = index_plan["paired"]
    spot_token = index_plan["spot_token"]
    spot_price = last_price_from_state(latest_ticks, spot_token)

    atm_strike: int | float | None = None
    atm_position: int | None = None
    if spot_price is not None and not paired.empty:
        atm_position = (paired["strike"] - spot_price).abs().idxmin()
        atm_strike = display_number(paired.loc[atm_position, "strike"])

    atm_straddle: float | None = None
    for _, row in paired.iterrows():
        strike_output = display_number(row["strike"])
        ce_token = int(row["ce_token"])
        pe_token = int(row["pe_token"])
        call_ltp = last_price_from_state(latest_ticks, ce_token)
        put_ltp = last_price_from_state(latest_ticks, pe_token)
        straddle = None
        if call_ltp is not None or put_ltp is not None:
            straddle = round((call_ltp or 0.0) + (put_ltp or 0.0), 2)

        atm_marker = "ATM" if atm_strike is not None and strike_output == atm_strike else ""
        if atm_marker == "ATM":
            atm_straddle = straddle

    status_prefix = "Streaming live ticks" if spot_price is not None else "Waiting for spot ticks"
    return {
        "expiry": index_plan["expiry"],
        "spot_price": round(spot_price, 2) if spot_price is not None else None,
        "atm_strike": atm_strike,
        "atm_position": int(atm_position) if atm_position is not None else None,
        "atm_straddle": atm_straddle,
        "status": f"{status_prefix} for {index_name} {index_plan['expiry']}",
    }


def write_index_header(index_sheet: xw.Sheet, result: dict[str, Any]) -> None:
    index_sheet.range("B3:E5").value = [
        [result["expiry"], None, "Spot", result["spot_price"]],
        [result["atm_strike"], None, "ATM Straddle", result["atm_straddle"]],
        [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None, "Status", result["status"]],
    ]


def render_indices(
    app: xw.App,
    workbook: xw.Book,
    plan: dict[str, Any],
    latest_ticks: dict[int, dict[str, Any]],
    changed_tokens: set[int],
    dirty_indices: set[str],
) -> None:
    summary_sheet = workbook.sheets["Summary"]
    raw_sheet = workbook.sheets["RawData"]
    try:
        app.api.ScreenUpdating = False
        write_raw_updates(raw_sheet, plan, latest_ticks, changed_tokens)
        for index_name in INDEX_ORDER:
            if index_name not in dirty_indices:
                continue
            index_sheet = workbook.sheets[index_name]
            result = build_index_result(index_name, plan, latest_ticks)
            write_index_header(index_sheet, result)
            index_plan = plan["indices"][index_name]
            current_anchor = index_plan.get("last_window_anchor")
            desired_anchor = None
            if result["atm_position"] is not None:
                desired_anchor, _ = build_display_window(index_plan, result["atm_position"])
            if desired_anchor != current_anchor:
                window_anchor = write_index_window(index_sheet, index_plan, result["atm_position"])
                index_plan["last_window_anchor"] = window_anchor
        summary_sheet.range("B9").value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    finally:
        app.api.ScreenUpdating = True


def write_error_summary(workbook: xw.Book, message: str) -> None:
    summary_sheet = workbook.sheets["Summary"]
    summary_sheet.range("B9").value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row_number in range(4, 7):
        summary_sheet.range(f"F{row_number}").value = message
    for index_name in INDEX_ORDER:
        workbook.sheets[index_name].range("E5").value = message


def run_once_render(
    app: xw.App,
    workbook: xw.Book,
    kite: KiteConnect,
    plan: dict[str, Any],
) -> None:
    latest_ticks = seed_initial_ticks(kite, plan)
    changed_tokens = set(plan["subscription_tokens"])
    dirty_indices = set(INDEX_ORDER)
    render_indices(app, workbook, plan, latest_ticks, changed_tokens, dirty_indices)
    workbook.save()
    log("Loaded a one-time snapshot into simple_option_chain.xlsx.")


def run_tick_stream(
    app: xw.App,
    workbook: xw.Book,
    kite: KiteConnect,
    access_token: str,
    plan: dict[str, Any],
) -> None:
    latest_ticks = seed_initial_ticks(kite, plan)
    render_indices(app, workbook, plan, latest_ticks, set(plan["subscription_tokens"]), set(INDEX_ORDER))
    workbook.save()
    log("Initial snapshot loaded into workbook.")

    tick_queue: queue.Queue[list[dict[str, Any]]] = queue.Queue()
    connected_event = threading.Event()
    closed_event = threading.Event()
    error_holder: dict[str, str] = {}

    ticker = KiteTicker(
        config.cred["api_key"],
        access_token,
        reconnect=True,
        reconnect_max_tries=50,
        reconnect_max_delay=60,
    )

    def on_connect(ws, response):
        log("WebSocket connected. Subscribing to spot and option tokens.")
        ws.subscribe(plan["subscription_tokens"])
        if plan["spot_tokens"]:
            ws.set_mode(ws.MODE_QUOTE, plan["spot_tokens"])
        if plan["option_tokens"]:
            ws.set_mode(ws.MODE_FULL, plan["option_tokens"])
        connected_event.set()

    def on_ticks(ws, ticks):
        if ticks:
            tick_queue.put(ticks)

    def on_close(ws, code, reason):
        log(f"WebSocket closed: {code} {reason}")
        closed_event.set()

    def on_error(ws, code, reason):
        error_holder["message"] = f"WebSocket error {code}: {reason}"
        log(error_holder["message"])

    ticker.on_connect = on_connect
    ticker.on_ticks = on_ticks
    ticker.on_close = on_close
    ticker.on_error = on_error

    log("Starting tick-by-tick WebSocket stream.")
    ticker.connect(threaded=True)
    if not connected_event.wait(timeout=20):
        raise RuntimeError("WebSocket did not connect within 20 seconds.")

    last_save_time = time.monotonic()
    first_live_tick_logged = False
    last_progress_log = time.monotonic()
    pending_changed_tokens: set[int] = set()
    pending_dirty_indices: set[str] = set()
    next_render_time = time.monotonic() + BATCH_RENDER_SECONDS
    try:
        while True:
            try:
                timeout = max(0.01, min(QUEUE_POLL_SECONDS, next_render_time - time.monotonic()))
                ticks_batch = tick_queue.get(timeout=timeout)
            except queue.Empty:
                ticks_batch = None

            if ticks_batch is not None:
                queued_batches = [ticks_batch]
                while True:
                    try:
                        queued_batches.append(tick_queue.get_nowait())
                    except queue.Empty:
                        break

                for batch in queued_batches:
                    for tick in batch:
                        token = int(tick["instrument_token"])
                        latest_ticks[token] = tick
                        pending_changed_tokens.add(token)
                        if token in plan["option_token_to_index"]:
                            pending_dirty_indices.add(plan["option_token_to_index"][token])
                        if token in plan["spot_token_to_index"]:
                            pending_dirty_indices.add(plan["spot_token_to_index"][token])

            now = time.monotonic()
            if pending_changed_tokens and now >= next_render_time:
                render_indices(
                    app,
                    workbook,
                    plan,
                    latest_ticks,
                    set(pending_changed_tokens),
                    set(pending_dirty_indices) or set(INDEX_ORDER),
                )
                if not first_live_tick_logged:
                    log("Live ticks received. Excel is streaming now.")
                    first_live_tick_logged = True
                    last_progress_log = now
                elif now - last_progress_log >= 5:
                    log(f"Streaming normally. Latest batch touched {len(pending_changed_tokens)} instruments.")
                    last_progress_log = now

                pending_changed_tokens.clear()
                pending_dirty_indices.clear()
                next_render_time = now + BATCH_RENDER_SECONDS
            elif now >= next_render_time:
                next_render_time = now + BATCH_RENDER_SECONDS

            if closed_event.is_set() and tick_queue.empty() and not pending_changed_tokens:
                break

            if now - last_save_time >= AUTOSAVE_SECONDS:
                workbook.save()
                last_save_time = now
    except KeyboardInterrupt:
        log("Updater stopped by user.")
    finally:
        try:
            ticker.close()
        except Exception:
            pass
        try:
            workbook.save()
        except Exception:
            pass


def default_expiries_from_frame(instrument_frame: pd.DataFrame) -> dict[str, str]:
    expiry_map = build_expiry_map(instrument_frame)
    return {index_name: nearest_expiry(expiry_map[index_name]) for index_name in INDEX_ORDER}


def main() -> None:
    args = parse_args()

    if args.rebuild_only:
        instrument_frame = load_instrument_universe(kite=None)
        ensure_workbook(default_expiries_from_frame(instrument_frame))
        log("Workbook rebuild completed.")
        return

    access_token = load_access_token()
    kite = build_kite_client(access_token)
    instrument_frame = load_instrument_universe(kite)
    ensure_workbook(default_expiries_from_frame(instrument_frame))
    plan = build_runtime_plan(kite, instrument_frame)

    app, workbook = open_excel_workbook()
    prepare_workbook_for_plan(workbook, plan)

    try:
        if args.once:
            run_once_render(app, workbook, kite, plan)
        else:
            run_tick_stream(app, workbook, kite, access_token, plan)
    except Exception as exc:
        message = f"Refresh failed: {exc}"
        write_error_summary(workbook, message)
        try:
            workbook.save()
        except Exception:
            pass
        raise


if __name__ == "__main__":
    main()
