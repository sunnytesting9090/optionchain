from __future__ import annotations

import argparse
import os
import shutil
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import xlwings as xw
from kiteconnect import KiteConnect
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

import config


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "option_chain.xlsx"
ACCESS_TOKEN_PATH = BASE_DIR / "access_token.txt"
CACHE_FILES = {
    "NFO": BASE_DIR / "instruments_nfo.csv",
    "BFO": BASE_DIR / "instruments_bfo.csv",
}

WORKBOOK_VERSION = "option-chain-v3"
REFRESH_INTERVAL_SECONDS = 0.5
CACHE_MAX_AGE_HOURS = 12
STRIKES_EACH_SIDE = 10
HEADER_ROW = 11
DATA_START_ROW = 12
MAX_DISPLAY_ROWS = (STRIKES_EACH_SIDE * 2) + 1

SUPPORTED_INDICES = {
    "NIFTY": {
        "name": "NIFTY",
        "segment": "NFO-OPT",
        "exchange": "NFO",
        "spot_symbol": "NSE:NIFTY 50",
    },
    "BANKNIFTY": {
        "name": "BANKNIFTY",
        "segment": "NFO-OPT",
        "exchange": "NFO",
        "spot_symbol": "NSE:NIFTY BANK",
    },
    "SENSEX": {
        "name": "SENSEX",
        "segment": "BFO-OPT",
        "exchange": "BFO",
        "spot_symbol": "BSE:SENSEX",
    },
}
INDEX_ORDER = list(SUPPORTED_INDICES.keys())


def log(message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}", flush=True)


def log_step(step_number: int, message: str) -> None:
    log(f"Step {step_number}: {message}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Live option chain updater for NIFTY, BANKNIFTY, and SENSEX."
    )
    parser.add_argument(
        "--once",
        action="store_true",
        help="Refresh the Excel sheet once and exit.",
    )
    parser.add_argument(
        "--rebuild-only",
        action="store_true",
        help="Rebuild the workbook template and exit without starting the live loop.",
    )
    parser.add_argument(
        "--interval",
        type=int,
        default=REFRESH_INTERVAL_SECONDS,
        help="Refresh interval in seconds for live mode.",
    )
    return parser.parse_args()


def load_access_token() -> str:
    if not ACCESS_TOKEN_PATH.exists():
        raise FileNotFoundError(
            "access_token.txt is missing. Run run_option_chain.bat and paste today's access token."
        )

    access_token = ACCESS_TOKEN_PATH.read_text(encoding="utf-8").strip()
    if not access_token:
        raise ValueError(
            "access_token.txt is empty. Run run_option_chain.bat and paste today's access token."
        )

    return access_token


def build_kite_client() -> KiteConnect:
    log_step(1, "Loading API credentials from config.py.")
    kite = KiteConnect(api_key=config.cred["api_key"])

    log_step(2, "Loading the access token from access_token.txt.")
    kite.set_access_token(load_access_token())

    profile = kite.profile()
    user_name = profile.get("user_name") or profile.get("user_id") or "Unknown user"
    log(f"Connected to Kite as {user_name}.")
    return kite


def cache_is_stale(cache_path: Path) -> bool:
    if not cache_path.exists():
        return True

    age_seconds = time.time() - cache_path.stat().st_mtime
    return age_seconds > (CACHE_MAX_AGE_HOURS * 3600)


def refresh_exchange_cache(kite: KiteConnect, exchange: str) -> None:
    cache_path = CACHE_FILES[exchange]
    if not cache_is_stale(cache_path):
        log(f"Using cached {exchange} instruments from {cache_path.name}.")
        return

    log_step(3, f"Refreshing the {exchange} instrument cache from Kite.")
    try:
        instruments = kite.instruments(exchange)
        pd.DataFrame(instruments).to_csv(cache_path, index=False)
        log(f"Saved {len(instruments)} rows to {cache_path.name}.")
    except Exception:
        if cache_path.exists():
            log(f"Could not refresh {exchange} right now, so the existing cache will be used.")
            return
        raise


def load_instrument_universe(kite: KiteConnect) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for exchange in sorted({meta["exchange"] for meta in SUPPORTED_INDICES.values()}):
        refresh_exchange_cache(kite, exchange)
        frame = pd.read_csv(CACHE_FILES[exchange], low_memory=False)
        frame["expiry"] = frame["expiry"].astype(str).str[:10]
        frames.append(frame)

    combined = pd.concat(frames, ignore_index=True)
    combined = combined[combined["segment"].isin({"NFO-OPT", "BFO-OPT"})].copy()
    combined["instrument_type"] = combined["instrument_type"].astype(str)
    combined["name"] = combined["name"].astype(str)
    return combined


def build_expiry_map(instrument_frame: pd.DataFrame) -> dict[str, list[str]]:
    expiry_map: dict[str, list[str]] = {}
    for index_name, meta in SUPPORTED_INDICES.items():
        filtered = instrument_frame[
            (instrument_frame["name"] == meta["name"])
            & (instrument_frame["segment"] == meta["segment"])
            & (instrument_frame["instrument_type"].isin(["CE", "PE"]))
        ]
        expiries = sorted(filtered["expiry"].dropna().astype(str).unique().tolist())
        if not expiries:
            raise RuntimeError(f"No expiries were found for {index_name}.")
        expiry_map[index_name] = expiries
    return expiry_map


def backup_existing_workbook() -> None:
    if not WORKBOOK_PATH.exists():
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = WORKBOOK_PATH.with_name(f"option_chain_backup_{timestamp}.xlsx")
    shutil.copy2(WORKBOOK_PATH, backup_path)
    log(f"Existing workbook backed up to {backup_path.name}.")


def workbook_requires_rebuild() -> bool:
    if not WORKBOOK_PATH.exists():
        return True

    try:
        workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
        try:
            if "OptionChain" not in workbook.sheetnames or "Lists" not in workbook.sheetnames:
                return True
            return workbook["Lists"]["Z1"].value != WORKBOOK_VERSION
        finally:
            workbook.close()
    except Exception:
        return True


def build_inline_dropdown_formula(values: list[str]) -> str:
    cleaned_values = [str(value).strip() for value in values if str(value).strip()]
    return '"' + ",".join(cleaned_values) + '"'


def create_workbook_template(expiry_map: dict[str, list[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OptionChain"
    lists_sheet = workbook.create_sheet("Lists")

    sheet.merge_cells("A1:H1")
    sheet["A1"] = "Live Option Chain"
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")

    sheet.merge_cells("A2:H2")
    sheet["A2"] = "Change the index and expiry from the dropdowns. The Python updater refreshes the data automatically."
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"].font = Font(italic=True, color="404040")
    sheet["A2"].fill = PatternFill("solid", fgColor="D9EAF7")

    label_fill = PatternFill("solid", fgColor="E2F0D9")
    summary_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_border = Border(
        left=Side(style="thin", color="C9C9C9"),
        right=Side(style="thin", color="C9C9C9"),
        top=Side(style="thin", color="C9C9C9"),
        bottom=Side(style="thin", color="C9C9C9"),
    )

    control_cells = {
        "A4": "Index",
        "D4": "Expiry",
        "G4": "Spot",
        "G5": "ATM Strike",
        "G6": "ATM Straddle Sum",
        "G7": "Displayed Straddle Total",
        "G8": "Last Updated",
        "G9": "Status",
    }
    for cell, value in control_cells.items():
        sheet[cell] = value
        sheet[cell].font = Font(bold=True)
        sheet[cell].fill = label_fill if cell in {"A4", "D4"} else summary_fill
        sheet[cell].border = thin_border

    sheet["B4"] = INDEX_ORDER[0]
    sheet["E4"] = expiry_map[INDEX_ORDER[0]][0]
    for cell in ("B4", "E4", "H4", "H5", "H6", "H7", "H8", "H9"):
        sheet[cell].border = thin_border
    sheet["H9"] = "Waiting for first refresh..."

    headers = ["ATM", "Strike", "Call LTP", "Put LTP", "Straddle Sum"]
    for column_number, header in enumerate(headers, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=column_number)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="548235")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_number in range(DATA_START_ROW, DATA_START_ROW + MAX_DISPLAY_ROWS):
        for column_number in range(1, 6):
            cell = sheet.cell(row=row_number, column=column_number)
            cell.border = thin_border
            if column_number == 2:
                cell.number_format = "0"
            elif column_number >= 3:
                cell.number_format = "0.00"

    for cell in ("H4", "H5"):
        sheet[cell].number_format = "0"
    for cell in ("H6", "H7"):
        sheet[cell].number_format = "0.00"

    sheet.freeze_panes = f"A{DATA_START_ROW}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 20
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 14
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["G"].width = 22
    sheet.column_dimensions["H"].width = 24

    index_validation = DataValidation(
        type="list",
        formula1=build_inline_dropdown_formula(INDEX_ORDER),
        allow_blank=False,
        showDropDown=True,
    )
    expiry_validation = DataValidation(
        type="list",
        formula1=build_inline_dropdown_formula(expiry_map[INDEX_ORDER[0]]),
        allow_blank=False,
        showDropDown=True,
    )
    sheet.add_data_validation(index_validation)
    sheet.add_data_validation(expiry_validation)
    index_validation.add(sheet["B4"])
    expiry_validation.add(sheet["E4"])

    highlight_fill = PatternFill("solid", fgColor="FFE699")
    sheet.conditional_formatting.add(
        f"A{DATA_START_ROW}:E{DATA_START_ROW + MAX_DISPLAY_ROWS - 1}",
        FormulaRule(formula=[f'$A{DATA_START_ROW}="ATM"'], fill=highlight_fill),
    )

    lists_sheet["A1"] = "WorkbookVersion"
    lists_sheet["A2"] = WORKBOOK_VERSION
    lists_sheet.sheet_state = "hidden"
    lists_sheet["Z1"] = WORKBOOK_VERSION

    workbook.save(WORKBOOK_PATH)


def ensure_workbook(expiry_map: dict[str, list[str]]) -> None:
    if workbook_requires_rebuild():
        if WORKBOOK_PATH.exists():
            backup_existing_workbook()
        log_step(4, "Creating a fresh option_chain.xlsx workbook with dropdowns and formatting.")
        create_workbook_template(expiry_map)
        return

    log_step(4, "Existing workbook structure is valid, so it will be reused.")


def normalize_index_selection(value: Any) -> str:
    if value is None:
        return INDEX_ORDER[0]

    selected = str(value).strip().upper()
    return selected if selected in SUPPORTED_INDICES else INDEX_ORDER[0]


def normalize_expiry_selection(value: Any) -> str | None:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()
    if not text:
        return None
    return text[:10]


def find_open_workbook() -> tuple[xw.App | None, xw.Book | None]:
    workbook_path = os.path.normcase(str(WORKBOOK_PATH.resolve()))

    for app in xw.apps:
        for book in app.books:
            try:
                if os.path.normcase(book.fullname) == workbook_path:
                    return app, book
            except Exception:
                continue

    return None, None


def open_excel_workbook() -> tuple[xw.App, xw.Book]:
    existing_app, existing_book = find_open_workbook()
    if existing_app and existing_book:
        log("Using the workbook that is already open in Excel.")
        return existing_app, existing_book

    log_step(5, "Opening the workbook in Excel.")
    app = xw.App(visible=True, add_book=False)
    app.display_alerts = False
    book = app.books.open(str(WORKBOOK_PATH))
    book.activate(steal_focus=True)
    return app, book


def apply_dropdown_validation(target_range: xw.Range, values: list[str]) -> None:
    formula = build_inline_dropdown_formula(values)
    validation = target_range.api.Validation
    try:
        validation.Delete()
    except Exception:
        pass

    validation.Add(Type=3, AlertStyle=1, Operator=1, Formula1=formula)
    validation.IgnoreBlank = False
    validation.InCellDropdown = True


def sync_dropdowns(
    option_sheet: xw.Sheet, expiry_map: dict[str, list[str]]
) -> tuple[str, str]:
    apply_dropdown_validation(option_sheet.range("B4"), INDEX_ORDER)

    selected_index = normalize_index_selection(option_sheet.range("B4").value)
    available_expiries = expiry_map[selected_index]

    apply_dropdown_validation(option_sheet.range("E4"), available_expiries)

    selected_expiry = normalize_expiry_selection(option_sheet.range("E4").value)
    if selected_expiry not in available_expiries:
        selected_expiry = available_expiries[0]
        option_sheet.range("E4").value = selected_expiry

    option_sheet.range("B4").value = selected_index
    return selected_index, selected_expiry


def ensure_selection_controls(
    option_sheet: xw.Sheet, expiry_map: dict[str, list[str]]
) -> tuple[str, str]:
    selected_index = normalize_index_selection(option_sheet.range("B4").value)
    available_expiries = expiry_map[selected_index]

    selected_expiry = normalize_expiry_selection(option_sheet.range("E4").value)
    if selected_expiry not in available_expiries:
        selected_expiry = available_expiries[0]
        option_sheet.range("E4").value = selected_expiry

    option_sheet.range("B4").value = selected_index
    return selected_index, selected_expiry


def fetch_option_chain(
    kite: KiteConnect,
    instrument_frame: pd.DataFrame,
    selected_index: str,
    selected_expiry: str,
) -> dict[str, Any]:
    meta = SUPPORTED_INDICES[selected_index]
    filtered = instrument_frame[
        (instrument_frame["name"] == meta["name"])
        & (instrument_frame["segment"] == meta["segment"])
        & (instrument_frame["expiry"] == selected_expiry)
        & (instrument_frame["instrument_type"].isin(["CE", "PE"]))
    ].copy()

    if filtered.empty:
        raise RuntimeError(f"No option contracts were found for {selected_index} {selected_expiry}.")

    ce_frame = (
        filtered[filtered["instrument_type"] == "CE"][["strike", "instrument_token"]]
        .drop_duplicates("strike")
        .rename(columns={"instrument_token": "ce_token"})
    )
    pe_frame = (
        filtered[filtered["instrument_type"] == "PE"][["strike", "instrument_token"]]
        .drop_duplicates("strike")
        .rename(columns={"instrument_token": "pe_token"})
    )

    paired = ce_frame.merge(pe_frame, on="strike", how="inner").sort_values("strike").reset_index(drop=True)
    if paired.empty:
        raise RuntimeError(f"No CE/PE strike pairs are available for {selected_index} {selected_expiry}.")

    spot_quote = kite.ltp([meta["spot_symbol"]])
    spot_price = float(spot_quote[meta["spot_symbol"]]["last_price"])

    atm_position = (paired["strike"] - spot_price).abs().idxmin()
    atm_strike = float(paired.loc[atm_position, "strike"])

    start_position = max(atm_position - STRIKES_EACH_SIDE, 0)
    end_position = min(atm_position + STRIKES_EACH_SIDE + 1, len(paired))
    display_frame = paired.iloc[start_position:end_position].copy()

    tokens = display_frame["ce_token"].tolist() + display_frame["pe_token"].tolist()
    option_quotes = kite.ltp(tokens)

    def last_price_for(token: int) -> float:
        quote = option_quotes.get(token) or option_quotes.get(str(token)) or {}
        return float(quote.get("last_price", 0.0) or 0.0)

    display_frame["call_ltp"] = display_frame["ce_token"].map(last_price_for)
    display_frame["put_ltp"] = display_frame["pe_token"].map(last_price_for)
    display_frame["straddle_sum"] = display_frame["call_ltp"] + display_frame["put_ltp"]
    display_frame["atm_marker"] = display_frame["strike"].apply(
        lambda strike: "ATM" if float(strike) == atm_strike else ""
    )

    atm_row = display_frame[display_frame["atm_marker"] == "ATM"]
    atm_straddle_sum = (
        float(atm_row.iloc[0]["straddle_sum"])
        if not atm_row.empty
        else float(display_frame.iloc[(display_frame["strike"] - atm_strike).abs().idxmin()]["straddle_sum"])
    )
    displayed_total = float(display_frame["straddle_sum"].sum())

    rows: list[list[Any]] = []
    for _, row in display_frame.iterrows():
        strike_value = float(row["strike"])
        strike_output = int(strike_value) if strike_value.is_integer() else round(strike_value, 2)
        rows.append(
            [
                row["atm_marker"],
                strike_output,
                round(float(row["call_ltp"]), 2),
                round(float(row["put_ltp"]), 2),
                round(float(row["straddle_sum"]), 2),
            ]
        )

    return {
        "index": selected_index,
        "expiry": selected_expiry,
        "spot_price": round(spot_price, 2),
        "atm_strike": int(atm_strike) if atm_strike.is_integer() else round(atm_strike, 2),
        "atm_straddle_sum": round(atm_straddle_sum, 2),
        "displayed_total": round(displayed_total, 2),
        "rows": rows,
    }


def clear_existing_rows(option_sheet: xw.Sheet) -> None:
    option_sheet.range(
        (DATA_START_ROW, 1),
        (DATA_START_ROW + MAX_DISPLAY_ROWS - 1, 5),
    ).clear_contents()


def write_chain_to_sheet(option_sheet: xw.Sheet, chain_data: dict[str, Any]) -> None:
    clear_existing_rows(option_sheet)

    if chain_data["rows"]:
        option_sheet.range((DATA_START_ROW, 1)).value = chain_data["rows"]

    option_sheet.range("H4").value = chain_data["spot_price"]
    option_sheet.range("H5").value = chain_data["atm_strike"]
    option_sheet.range("H6").value = chain_data["atm_straddle_sum"]
    option_sheet.range("H7").value = chain_data["displayed_total"]
    option_sheet.range("H8").value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    option_sheet.range("H9").value = f"Live data loaded for {chain_data['index']} {chain_data['expiry']}"


def write_error_to_sheet(option_sheet: xw.Sheet, error_message: str) -> None:
    option_sheet.range("H8").value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    option_sheet.range("H9").value = error_message


def run_live_loop(
    kite: KiteConnect,
    instrument_frame: pd.DataFrame,
    expiry_map: dict[str, list[str]],
    interval_seconds: int,
    run_once: bool,
) -> None:
    _, workbook = open_excel_workbook()
    option_sheet = workbook.sheets["OptionChain"]
    workbook.save()

    log_step(6, "Starting the live refresh loop.")
    while True:
        try:
            selected_index, selected_expiry = sync_dropdowns(option_sheet, expiry_map)
            chain_data = fetch_option_chain(kite, instrument_frame, selected_index, selected_expiry)
            write_chain_to_sheet(option_sheet, chain_data)
            workbook.save()
            log(
                f"Updated {selected_index} {selected_expiry} with {len(chain_data['rows'])} strikes around ATM."
            )
        except KeyboardInterrupt:
            write_error_to_sheet(option_sheet, "Updater stopped by user.")
            workbook.save()
            log("Updater stopped by user.")
            raise
        except Exception as exc:
            error_message = f"Refresh failed: {exc}"
            write_error_to_sheet(option_sheet, error_message)
            workbook.save()
            log(error_message)

        if run_once:
            return

        time.sleep(max(interval_seconds, 2))


def main() -> None:
    args = parse_args()
    kite = build_kite_client()
    instrument_frame = load_instrument_universe(kite)
    expiry_map = build_expiry_map(instrument_frame)
    ensure_workbook(expiry_map)

    if args.rebuild_only:
        log("Workbook rebuild completed.")
        return

    run_live_loop(
        kite=kite,
        instrument_frame=instrument_frame,
        expiry_map=expiry_map,
        interval_seconds=args.interval,
        run_once=args.once,
    )


if __name__ == "__main__":
    main()
